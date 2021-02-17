# -*- coding: utf-8 -*-
"""
Created on Tue Feb 16 17:52:31 2021

@author: Francesco Renzi

target: creare n cartelle di lavoro usando un template dato:

1. python: modulo openpyxl -> carica template, lo attiva e lo popola, reiterato n volte con ciclo for 
https://openpyxl.readthedocs.io/en/stable/tutorial.html

#shebang py in caso di esecuzione su macchina linux
"""



import os, fnmatch

# setta la working directory
DIR = input('Inserisci il path del template Excel da usare: ')

def find(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
    return result

#occhio metti un break se no non ne esci più XD
while False:
    try:
        os.chdir(DIR)
        if find('*Template.xlsx', DIR):
            print("\nLa working directory è la seguente: {0}".format(os.getcwd()))
        else:
            raise FileNotFoundError
    except FileNotFoundError:
        print("\nIl template non è in questa directory")
    except NotADirectoryError:
        print("\n{0} non è una directory".format(DIR))
    except PermissionError:
        print("\nNon hai i permessi per accedere a {0}".format(DIR))
#%%


from openpyxl import load_workbook
   
wb = load_workbook('template_SAC.xlsx')
   
# modalita modifica
ws = wb.active
   
# metto valore nelle celle
ws['A2'] = 'si'
ws['B2'] = 23
   
ws['A3'] = 'no'
ws['B3'] = 'http://cmtool[...]'
   
# salvo il file
wb.save("<path>\<variabile_nome_file>")

